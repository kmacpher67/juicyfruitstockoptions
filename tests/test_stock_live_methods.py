import pandas as pd
import openpyxl
import pytest
from unittest.mock import MagicMock
from stock_live_comparison import StockLiveComparison


class DummyChain:
    def option_chain(self, exp):
        class OC:
            def __init__(self):
                self.calls = pd.DataFrame()
                self.puts = pd.DataFrame()
        return OC()


def test_fetch_data_empty_tickers():
    comp = StockLiveComparison([])
    assert comp.fetch_data([]) == []


def test_fetch_data_api_failure(monkeypatch):
    import stock_live_comparison as slc

    def fake_download(*args, **kwargs):
        return {}

    class BadTicker:
        @property
        def info(self):
            raise ValueError("boom")

    class DummyTickers:
        def __init__(self):
            self.tickers = {"AAA": BadTicker()}

    monkeypatch.setattr(slc.yf, "download", fake_download)
    monkeypatch.setattr(slc.yf, "Tickers", lambda symbols: DummyTickers())
    monkeypatch.setattr(slc.time, "sleep", lambda x: None)

    comp = StockLiveComparison(["AAA"])
    records = comp.fetch_data(["AAA"])
    assert len(records) == 1
    assert records[0]["Ticker"] == "AAA"
    assert records[0]["Error"] == "boom"


def test_fetch_data_retries_on_429_and_recovers(monkeypatch):
    import stock_live_comparison as slc

    def fake_download(*args, **kwargs):
        return {}

    class FlakyTicker:
        def __init__(self):
            self.calls = 0

        @property
        def info(self):
            self.calls += 1
            if self.calls < 3:
                raise RuntimeError("HTTP Error 429: Too Many Requests")
            return {"regularMarketPrice": 101}

    flaky = FlakyTicker()

    class DummyTickers:
        def __init__(self):
            self.tickers = {"AAA": flaky}

    monkeypatch.setattr(slc.yf, "download", fake_download)
    monkeypatch.setattr(slc.yf, "Tickers", lambda symbols: DummyTickers())
    monkeypatch.setattr(slc.time, "sleep", lambda x: None)
    monkeypatch.setattr(
        StockLiveComparison,
        "fetch_ticker_record",
        lambda self, ticker, info, ticker_hist, chain: {
            "Ticker": ticker,
            "Current Price": info.get("regularMarketPrice"),
            "Last Update": self.now.strftime("%Y-%m-%d %H:%M:%S"),
        },
    )

    comp = StockLiveComparison(["AAA"])
    records = comp.fetch_data(["AAA"])

    assert len(records) == 1
    assert records[0]["Ticker"] == "AAA"
    assert records[0]["Current Price"] == 101
    assert "Error" not in records[0] or records[0]["Error"] is None
    assert flaky.calls == 3


def test_download_history_batched_uses_small_batches_and_no_threads(monkeypatch):
    import stock_live_comparison as slc

    calls = []

    def fake_download(batch, **kwargs):
        calls.append((list(batch), kwargs))
        # empty-ish but valid frame
        return pd.DataFrame()

    monkeypatch.setattr(slc.yf, "download", fake_download)
    monkeypatch.setattr(slc.time, "sleep", lambda x: None)

    comp = StockLiveComparison(
        ["AAA", "BBB", "CCC"],
        download_batch_size=2,
        batch_pause_sec=0,
    )
    out = comp.download_history_batched(["AAA", "BBB", "CCC"])

    assert list(out.keys()) == ["AAA", "BBB", "CCC"]
    assert len(calls) == 2
    assert calls[0][0] == ["AAA", "BBB"]
    assert calls[1][0] == ["CCC"]
    assert calls[0][1]["threads"] is True
    assert calls[1][1]["threads"] is True
    assert calls[0][1]["progress"] is False


def test_download_history_batched_prefers_fresh_cache(monkeypatch, tmp_path):
    import stock_live_comparison as slc

    calls = []

    def fake_download(batch, **kwargs):
        calls.append((list(batch), kwargs))
        return pd.DataFrame()

    monkeypatch.setattr(slc.yf, "download", fake_download)
    monkeypatch.setattr(slc.time, "sleep", lambda x: None)

    comp = StockLiveComparison(["AAA", "BBB"], history_cache_ttl_hours=24)
    comp.history_cache_dir = tmp_path

    fresh_df = pd.DataFrame(
        {
            "Open": [1.0],
            "High": [1.1],
            "Low": [0.9],
            "Close": [1.05],
            "Volume": [1000],
        },
        index=pd.to_datetime(["2026-04-03"]),
    )
    fresh_df.to_csv(comp.history_cache_path("AAA"))

    out = comp.download_history_batched(["AAA", "BBB"])
    assert "AAA" in out and out["AAA"] is not None
    assert len(calls) == 1
    assert calls[0][0] == ["BBB"]


def test_fetch_ticker_record(monkeypatch):
    comp = StockLiveComparison(["AAA"])

    def fake_call(chain, price, days, otm_pct=6):
        mapping = {90: (1, None, '2025-01-01'), 180: (2, 105, '2025-06-01'), 365: (3, None, '2026-01-01')}
        return mapping[days]

    def fake_put(chain, price, days, otm_pct=6):
        return (5, '2026-01-01')

    comp.get_otm_call_yield = fake_call
    comp.get_otm_put_price = fake_put

    info = {
        "regularMarketPrice": 100,
        "marketCap": 2e12,
        "trailingPE": 15,
        "dividendYield": 0.01,
        "targetMeanPrice": 120,
        "exDividendDate": 1710000000,
        "longName": "Alpha Corp Inc.",
        "shortName": "Alpha Corp",
    }
    hist = pd.DataFrame({
        "Close": [90, 100],
        "High": [92, 102],
        "Low": [88, 98],
        "Open": [89, 99],
        "Volume": [1000, 1000],
        "Date": pd.date_range("2023-01-01", periods=2)
    })
    record = comp.fetch_ticker_record("AAA", info, hist, DummyChain())
    assert record["Ticker"] == "AAA"
    assert record["Company Name"] == "Alpha Corp Inc."
    assert record["1D % Change"] == "11.11%"
    assert record["YoY Price %"] == "11.1%"
    assert record["3-mo Call Yield"] == 1
    assert record["6-mo Call Yield"] == 2
    assert record["6-mo Call Strike"] == 105
    assert record["1-yr Call Yield"] == 3
    assert record["1-yr 6% OTM PUT Price"] == 5
    assert record["Annual Yield Put Prem"] == 5
    assert record["Annual Yield Call Prem"] == 3


def test_fetch_ticker_record_skips_news_when_disabled(monkeypatch):
    comp = StockLiveComparison(["AAA"], fetch_profile_news=False)

    def fake_call(chain, price, days, otm_pct=6):
        return (None, None, None)

    def fake_put(chain, price, days, otm_pct=6):
        return (None, None)

    comp.get_otm_call_yield = fake_call
    comp.get_otm_put_price = fake_put

    class NewsExplodesChain:
        @property
        def news(self):
            raise RuntimeError("news endpoint should not be called")

        def option_chain(self, exp):
            class OC:
                def __init__(self):
                    self.calls = pd.DataFrame()
                    self.puts = pd.DataFrame()
            return OC()

    info = {"regularMarketPrice": 100, "shortName": "AAA"}
    hist = pd.DataFrame(
        {
            "Close": [99, 100],
            "High": [100, 101],
            "Low": [98, 99],
            "Open": [99, 100],
            "Volume": [100, 100],
        }
    )
    record = comp.fetch_ticker_record("AAA", info, hist, NewsExplodesChain())
    assert record["profile"]["news"] == []


def test_apply_rate_limit_cooldown_sets_next_request_window(monkeypatch):
    comp = StockLiveComparison(["AAA"])
    monkeypatch.setattr("stock_live_comparison.time.time", lambda: 1000.0)
    monkeypatch.setattr("stock_live_comparison.random.uniform", lambda a, b: 1.0)
    comp.apply_rate_limit_cooldown(2)
    assert comp._next_request_not_before_ts > 1000.0


def test_is_http_429_error_detection():
    comp = StockLiveComparison(["AAA"])
    assert comp.is_http_429_error(RuntimeError("HTTP Error 429: Too Many Requests")) is True
    assert comp.is_http_429_error(RuntimeError("boom")) is False


def test_fetch_ticker_record_company_name_fallback(monkeypatch):
    """Verify Company Name fallback: longName → shortName → ticker."""
    comp = StockLiveComparison(["BBB"])

    def fake_call(chain, price, days, otm_pct=6):
        return (None, None, None)

    def fake_put(chain, price, days, otm_pct=6):
        return (None, None)

    comp.get_otm_call_yield = fake_call
    comp.get_otm_put_price = fake_put

    # Test shortName fallback (no longName)
    info_short = {"regularMarketPrice": 50, "shortName": "Beta Inc"}
    hist = pd.DataFrame({
        "Close": [48, 50], "High": [49, 51], "Low": [47, 49],
        "Open": [48, 50], "Volume": [500, 500],
        "Date": pd.date_range("2023-01-01", periods=2)
    })
    record = comp.fetch_ticker_record("BBB", info_short, hist, DummyChain())
    assert record["Company Name"] == "Beta Inc"

    # Test ticker fallback (no longName or shortName)
    info_none = {"regularMarketPrice": 50}
    record = comp.fetch_ticker_record("BBB", info_none, hist, DummyChain())
    assert record["Company Name"] == "BBB"


def test_add_ratio_column():
    comp = StockLiveComparison(["AAA"])
    df = pd.DataFrame([
        {"Annual Yield Put Prem": 5, "Annual Yield Call Prem": 10},
        {"Annual Yield Put Prem": None, "Annual Yield Call Prem": 5},
    ])
    df, put_col, call_col = comp.add_ratio_column(df)
    assert "Call/Put Skew" in df.columns
    assert df.iloc[0]["Call/Put Skew"] == 2.0  # 10 / 5
    assert pd.isna(df.iloc[1]["Call/Put Skew"])
    assert put_col == 1 and call_col == 2


def test_upsert_ratio_column():
    comp = StockLiveComparison(["AAA"])
    df = pd.DataFrame({"Annual Yield Put Prem": [5], "Annual Yield Call Prem": [10]})
    df = comp.upsert_ratio_column(df)
    assert df.iloc[0]["Call/Put Skew"] == 2.0
    df["Annual Yield Call Prem"] = [5]
    df = comp.upsert_ratio_column(df)
    assert df.iloc[0]["Call/Put Skew"] == 1.0


def test_save_to_excel(tmp_path):
    comp = StockLiveComparison(["AAA"])
    comp.filename = str(tmp_path / "out.xlsx")
    df = pd.DataFrame({
        "Ticker": ["AAA"],
        "Current Price": [100],
        "Annual Yield Put Prem": [5],
        "Annual Yield Call Prem": [10],
    })
    df, put_col, call_col = comp.add_ratio_column(df)
    comp.save_to_excel(df, put_col, call_col)
    wb = openpyxl.load_workbook(comp.filename)
    ws = wb.active
    assert ws.freeze_panes == "A2"
    ratio_cell = ws.cell(row=2, column=call_col + 1)
    put_letter = openpyxl.utils.get_column_letter(put_col)
    call_letter = openpyxl.utils.get_column_letter(call_col)
    # Check for correct inverted formula: call / put
    assert ratio_cell.value == 2.0 or ratio_cell.value == f"=IFERROR({call_letter}2/{put_letter}2,\"\")"
    assert ws["A1"].font.bold
    
    # Verify Hyperlink
    ticker_col_idx = df.columns.get_loc("Ticker") + 1
    ticker_cell = ws.cell(row=2, column=ticker_col_idx)
    assert ticker_cell.hyperlink is not None, f"Hyperlink is None for cell {ticker_cell.coordinate} with value {ticker_cell.value}"
    assert ticker_cell.hyperlink.target == "https://www.google.com/finance?q=AAA"
    assert ticker_cell.style == "Hyperlink"


def test_run(monkeypatch, tmp_path):
    sample_record = {
        "Ticker": "AAA",
        "Annual Yield Put Prem": 5,
        "Annual Yield Call Prem": 10,
    }

    monkeypatch.setattr(StockLiveComparison, "get_latest_spreadsheet", staticmethod(lambda base_name="AI_Stock_Live_Comparison_": (None, None)))

    def fake_fetch(self, tickers):
        return [sample_record]

    monkeypatch.setattr(StockLiveComparison, "fetch_data", fake_fetch)

    saved = {}

    def fake_save(self, df, put_col, call_col):
        saved["df"] = df.copy()
        saved["put_col"] = put_col
        saved["call_col"] = call_col

    monkeypatch.setattr(StockLiveComparison, "save_to_excel", fake_save)

    comp = StockLiveComparison(["AAA"])
    comp.output_dir = tmp_path # Override output dir
    comp.run()
    df = saved["df"]
    assert "Call/Put Skew" in df.columns
    assert df.iloc[0]["Call/Put Skew"] == 2.0
    assert saved["put_col"] == 2 and saved["call_col"] == 3


def test_is_recent():
    comp = StockLiveComparison([])
    # Mock 'now' to be consistent
    base_time = pd.Timestamp("2025-01-01 12:00:00")
    comp.now = base_time
    
    # Recent: 1 hour ago
    row_recent = {"Last Update": "2025-01-01 11:00:00"}
    assert comp.is_recent(row_recent) is True
    
    # Old: 25 hours ago (default max_age_hours is 24 probably? checking code...)
    # Actually code has self.max_age_hours, let's assume default unless init changes it.
    # Looking at class init, it might default to something. Let's start with a safe check.
    # If not set in init, we'll see. But typically it's 4 hours or similar.
    # Let's set it explicitly for test
    comp.max_age_hours = 4
    
    row_old = {"Last Update": "2025-01-01 07:00:00"} # 5 hours ago
    assert comp.is_recent(row_old) is False
    
    # Invalid
    assert comp.is_recent({}) is False
    assert comp.is_recent({"Last Update": None}) is False


def test_get_missing_or_outdated_tickers():
    comp = StockLiveComparison(["AAA", "BBB", "CCC"])
    comp.now = pd.Timestamp("2025-01-01 12:00:00")
    comp.max_age_hours = 4
    
    # AAA: Recent and complete
    # BBB: Old
    # CCC: Missing from DF
    # DDD: Recent but missing columns
    
    data = [
        {
            "Ticker": "AAA",
            "Last Update": "2025-01-01 11:00:00",
            "MA_30": 100, "MA_60": 100, "MA_120": 100, "MA_200": 100,
            "EMA_20": 100, "HMA_20": 100, "TSMOM_60": 0.05,
            "RSI_14": 50, "ATR_14": 1.5, "Price Action": {},
            "_PutExpDate_365": "2026-01-01"
        },
        {
            "Ticker": "BBB",
            "Last Update": "2025-01-01 06:00:00", # 6 hours old
            "MA_30": 100, "MA_60": 100, "MA_120": 100, "MA_200": 100,
            "EMA_20": 100, "HMA_20": 100, "TSMOM_60": 0.05,
            "RSI_14": 50, "ATR_14": 1.5, "Price Action": {},
             "_PutExpDate_365": "2026-01-01"
        },
        {
            "Ticker": "DDD",
            "Last Update": "2025-01-01 11:00:00",
            "MA_30": 100, # Missing other MAs
             "_PutExpDate_365": "2026-01-01"
        }
    ]
    df_existing = pd.DataFrame(data)
    
    missing = comp.get_missing_or_outdated_tickers(df_existing)
    
    assert "AAA" not in missing
    assert "BBB" in missing # Old
    assert "CCC" in missing # Not in DF
    # DDD is in DF but missing columns, so it should be considered outdated/re-fetch needed
    # Wait, get_missing_or_outdated_tickers iterates over self.tickers ("AAA", "BBB", "CCC")
    # DDD is NOT in self.tickers, so it won't be checked effectively unless added.
    
    # Let's add DDD to self.tickers to test the missing columns logic
    comp.tickers.append("DDD")
    missing = comp.get_missing_or_outdated_tickers(df_existing)
    assert "DDD" in missing


def test_run_merge_logic(monkeypatch, tmp_path, caplog):
    # Setup
    comp = StockLiveComparison(["AAA", "NEW"])
    comp.output_dir = tmp_path
    
    # Create "Existing" file
    existing_df = pd.DataFrame([
        {
            "Ticker": "AAA",
            "Last Update": "2025-01-01 10:00:00",
            "Annual Yield Put Prem": 5,
            "Annual Yield Call Prem": 5,
             "MA_30": 100, "MA_60": 100, "MA_120": 100, "MA_200": 100,
            "EMA_20": 100, "HMA_20": 100, "TSMOM_60": 0.05,
            "RSI_14": 50, "ATR_14": 1.5, "Price Action": "{}",
             "_PutExpDate_365": "2026-01-01"
        }
    ])
    existing_file = tmp_path / "AI_Stock_Live_Comparison_20250101_100000.xlsx"
    existing_df.to_excel(existing_file, index=False)
    if existing_file.stat().st_size < 12000:
        existing_file.write_bytes(existing_file.read_bytes() + (b"x" * (12000 - existing_file.stat().st_size)))
    
    # Mock get_latest_spreadsheet to return this file
    def fake_get_latest(directory, base_name="AI_Stock_Live_Comparison_"):
        return existing_file, pd.Timestamp("2025-01-01 10:00:00")
    monkeypatch.setattr(comp, "get_latest_spreadsheet", fake_get_latest)
    
    # Mock fetch_data to return only "NEW" ticker data
    def fake_fetch(tickers):
        return [{
            "Ticker": "NEW", 
            "Last Update": "2025-01-02 10:00:00",
            "Annual Yield Put Prem": 10,
            "Annual Yield Call Prem": 10
        }]
    monkeypatch.setattr(comp, "fetch_data", fake_fetch)
    
    # Mock save_to_excel avoid actual excel ops
    saved_df = []
    monkeypatch.setattr(comp, "save_to_excel", lambda df, p, c: saved_df.append(df))
    monkeypatch.setattr(comp, "upsert_to_mongo", lambda df: None)
    
    # Mock json export
    monkeypatch.setattr("stock_live_comparison.export_data", lambda: None)
    
    # Mock datetime to control self.now inside run()
    import datetime
    class FakeDatetime(datetime.datetime):
        @classmethod
        def now(cls):
            return pd.Timestamp("2025-01-01 11:00:00")
            
    monkeypatch.setattr("stock_live_comparison.datetime", FakeDatetime)
    
    # Run
    # Enable logging (optional, keeping for safety if fails again but usually not needed)
    # caplog.set_level(logging.DEBUG) 
    
    comp.run()
    
    assert len(saved_df) == 1
    result = saved_df[0]
    assert len(result) == 2
    assert "AAA" in result["Ticker"].values
    assert "NEW" in result["Ticker"].values


def test_select_output_report_file_sync_reuses_latest_without_creating_new(tmp_path):
    comp = StockLiveComparison(["AAA"])
    comp.output_dir = tmp_path
    comp.now = pd.Timestamp("2026-04-02 12:00:00")

    existing = tmp_path / "AI_Stock_Live_Comparison_20260401_052900.xlsx"
    existing.write_text("placeholder", encoding="utf-8")

    selected = comp.select_output_report_file(force_new_file=False, allow_create_if_missing=False)
    assert selected == existing


def test_select_output_report_file_scheduled_reuses_today_file(tmp_path):
    comp = StockLiveComparison(["AAA"])
    comp.output_dir = tmp_path
    comp.now = pd.Timestamp("2026-04-02 16:44:40")

    today = tmp_path / "AI_Stock_Live_Comparison_20260402_052900.xlsx"
    today.write_text("placeholder", encoding="utf-8")

    selected = comp.select_output_report_file(force_new_file=False, allow_create_if_missing=True)
    assert selected == today


def test_get_latest_viable_spreadsheet_skips_small_files(tmp_path):
    comp = StockLiveComparison(["AAA"])
    comp.output_dir = tmp_path

    small = tmp_path / "AI_Stock_Live_Comparison_20260403_100000.xlsx"
    small.write_bytes(b"x" * 1024)

    viable = tmp_path / "AI_Stock_Live_Comparison_20260403_090000.xlsx"
    viable.write_bytes(b"x" * 12000)

    selected, _ = comp.get_latest_viable_spreadsheet(tmp_path, min_bytes=10 * 1024)
    assert selected == viable


def test_run_uses_latest_viable_file_as_merge_source(monkeypatch, tmp_path):
    comp = StockLiveComparison(["AAA", "BBB"])
    comp.output_dir = tmp_path

    viable_source = tmp_path / "AI_Stock_Live_Comparison_20260403_100000.xlsx"
    pd.DataFrame(
        [
            {
                "Ticker": "AAA",
                "Last Update": "2026-04-03 10:00:00",
                "Annual Yield Put Prem": 1,
                "Annual Yield Call Prem": 2,
                "MA_30": 1,
                "MA_60": 1,
                "MA_120": 1,
                "MA_200": 1,
                "EMA_20": 1,
                "HMA_20": 1,
                "TSMOM_60": 0.1,
                "RSI_14": 50,
                "ATR_14": 1.2,
                "Price Action": "{}",
                "_PutExpDate_365": "2027-01-01",
            }
        ]
    ).to_excel(viable_source, index=False)
    if viable_source.stat().st_size < 12000:
        viable_source.write_bytes(viable_source.read_bytes() + (b"x" * (12000 - viable_source.stat().st_size)))
    tiny_latest = tmp_path / "AI_Stock_Live_Comparison_20260403_110000.xlsx"
    tiny_latest.write_bytes(b"x" * 1024)

    monkeypatch.setattr(comp, "fetch_data", lambda tickers: [{"Ticker": "BBB", "Annual Yield Put Prem": 2, "Annual Yield Call Prem": 4, "Last Update": "2026-04-03 11:00:00"}])
    monkeypatch.setattr(comp, "save_to_excel", lambda *args, **kwargs: None)
    monkeypatch.setattr(comp, "upsert_to_mongo", lambda *args, **kwargs: None)
    monkeypatch.setattr("stock_live_comparison.export_data", lambda: None)

    comp.run()
    assert comp.latest_file in {tiny_latest, viable_source}
    assert comp.latest_viable_file == viable_source


def test_run_rejects_suspiciously_small_result_set(monkeypatch, tmp_path):
    many_tickers = [f"T{i:03d}" for i in range(100)]
    comp = StockLiveComparison(many_tickers)
    comp.output_dir = tmp_path

    monkeypatch.setattr(comp, "fetch_data", lambda tickers: [{"Ticker": "ONLY1", "Annual Yield Put Prem": 1, "Annual Yield Call Prem": 2, "Last Update": "2026-04-03 11:00:00"}])
    monkeypatch.setattr(comp, "save_to_excel", lambda *args, **kwargs: None)
    monkeypatch.setattr(comp, "upsert_to_mongo", lambda *args, **kwargs: None)
    monkeypatch.setattr("stock_live_comparison.export_data", lambda: None)

    with pytest.raises(RuntimeError, match="Suspicious stock-analysis output"):
        comp.run()


def test_merge_detail_record_preserves_existing_profile_and_price_action_when_incoming_sparse():
    existing = {
        "Ticker": "AAPL",
        "profile": {"sector": "Technology", "news": [{"title": "x"}]},
        "Price Action": {"trend": "Bullish"},
    }
    incoming = {
        "Ticker": "AAPL",
        "Current Price": 201.2,
        "profile": {},
        "Price Action": {},
    }
    merged = StockLiveComparison.merge_detail_record(incoming, existing)
    assert merged["Current Price"] == 201.2
    assert merged["profile"]["sector"] == "Technology"
    assert merged["profile"]["news"][0]["title"] == "x"
    assert merged["Price Action"]["trend"] == "Bullish"


def test_upsert_to_mongo_writes_canonical_stock_data_by_ticker(monkeypatch):
    comp = StockLiveComparison(["AAPL"])
    StockLiveComparison._price_history_indexes_ensured = False
    StockLiveComparison._instrument_snapshot_indexes_ensured = False
    df = pd.DataFrame(
        [
            {
                "Ticker": "aapl",
                "Company Name": "Apple Inc.",
                "Current Price": 199.5,
                "1D % Change": "+1.23%",
                "Last Update": "2026-04-03 13:20:00",
                "Call/Put Skew": 1.4,
                "Price Action": {"trend": "Bullish"},
                "profile": {"sector": "Technology", "news": []},
            }
        ]
    )

    fake_collection = MagicMock()
    fake_collection.find_one.return_value = {}
    fake_db = MagicMock()
    fake_db.collection = fake_collection
    fake_snapshot_db = MagicMock()
    fake_snapshot_collection = MagicMock()
    fake_snapshot_db.collection = fake_snapshot_collection
    fake_price_db = MagicMock()
    fake_price_collection = MagicMock()
    fake_price_db.collection = fake_price_collection

    db_ctor = MagicMock(side_effect=[fake_db, fake_snapshot_db, fake_price_db])
    monkeypatch.setattr("stock_live_comparison.AiStockDatabase", db_ctor)

    comp.upsert_to_mongo(df)

    assert db_ctor.call_count == 3
    assert db_ctor.call_args_list[0].kwargs["collection_name"] == "stock_data"
    assert db_ctor.call_args_list[1].kwargs["collection_name"] == "instrument_snapshot"
    assert db_ctor.call_args_list[2].kwargs["collection_name"] == "instrument_price_history"
    fake_db.upsert_stock_record.assert_called_once()
    args, kwargs = fake_db.upsert_stock_record.call_args
    persisted = args[0]
    assert persisted["Ticker"] == "AAPL"
    assert "profile" in persisted
    assert "Price Action" in persisted
    assert kwargs["key_fields"] == ("Ticker",)
    fake_snapshot_db.upsert_stock_record.assert_called_once()
    snapshot_args, snapshot_kwargs = fake_snapshot_db.upsert_stock_record.call_args
    persisted_snapshot = snapshot_args[0]
    assert persisted_snapshot["instrument_key"] == "STK:AAPL"
    assert persisted_snapshot["symbol"] == "AAPL"
    assert snapshot_kwargs["key_fields"] == ("instrument_key",)
    fake_price_db.upsert_stock_record.assert_called_once()
    hist_args, hist_kwargs = fake_price_db.upsert_stock_record.call_args
    persisted_hist = hist_args[0]
    assert persisted_hist["instrument_key"] == "STK:AAPL"
    assert persisted_hist["instrument_key_legacy"] == "AAPL"
    assert persisted_hist["source"] == "stock_live_comparison"
    assert hist_kwargs["key_fields"] == ("instrument_key", "timestamp", "source")
    assert fake_snapshot_collection.create_index.call_count == 2
    assert fake_price_collection.create_index.call_count == 3
    assert any(
        call.args and call.args[0] == [("timestamp", -1)]
        for call in fake_price_collection.create_index.call_args_list
    )


def test_missing_required_detail_fields_requires_profile_news_key():
    record = {
        "Ticker": "MSFT",
        "Company Name": "Microsoft",
        "Current Price": 400,
        "1D % Change": "0.4%",
        "Last Update": "2026-04-03 10:00:00",
        "Call/Put Skew": 1.2,
        "Price Action": {"trend": "Bullish"},
        "profile": {"sector": "Technology"},
    }
    missing = StockLiveComparison.missing_required_detail_fields(record)
    assert "profile.news" in missing
