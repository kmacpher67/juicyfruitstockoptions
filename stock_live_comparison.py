from pathlib import Path
import re
import yfinance as yf
import pandas as pd
import time
import random
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from Ai_Stock_Database import AiStockDatabase
from export_mongo import export_data
from app.services.price_action_service import PriceActionService

class StockLiveComparison:
    """Collect stock metrics and export them to an Excel sheet."""
    _price_history_indexes_ensured = False

    def __init__(
        self,
        tickers,
        max_age_hours=4,
        highlight_threshold=0.05,
        fetch_profile_news=False,
        min_request_interval_sec=1.5,
        download_batch_size=6,
        batch_pause_sec=8.0,
        history_cache_ttl_hours=6.0,
    ):
        self.tickers = list(dict.fromkeys(tickers))
        self.max_age_hours = max_age_hours
        self.highlight_threshold = highlight_threshold  # e.g. 0.05 for 5%
        self.fetch_profile_news = bool(fetch_profile_news)
        self.min_request_interval_sec = float(min_request_interval_sec)
        self.download_batch_size = max(1, int(download_batch_size))
        self.batch_pause_sec = max(0.0, float(batch_pause_sec))
        self.history_cache_ttl_hours = max(0.0, float(history_cache_ttl_hours))
        self.records = []
        self.output_dir = Path("report-results")
        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True, exist_ok=True)
        self.history_cache_dir = Path("data/yf_history")
        self.history_cache_dir.mkdir(parents=True, exist_ok=True)
        self.now = datetime.now()
        self.filename = None
        self.latest_file = None
        self.latest_viable_file = None
        self.logger = logging.getLogger(__name__)
        self.min_viable_report_bytes = 10 * 1024
        self._next_request_not_before_ts = 0.0
        self._consecutive_429 = 0

    @staticmethod
    def setup_logging(log_file="stock_live_comparison.log"):
        """Configure logging to file (DEBUG) and console (INFO)."""
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)
        
        # Create formatters
        file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_formatter = logging.Formatter('%(message)s')
        
        # File Handler
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(file_formatter)
        logger.addHandler(file_handler)
        
        # Console Handler
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(console_formatter)
        logger.addHandler(console_handler)

    @staticmethod
    def calculate_moving_averages(hist, windows=(30, 60, 120, 200)):
        """Return a dict of moving averages for the given windows from a price series."""
        result = {f"MA_{w}": None for w in windows}
        if hist is None or hist.empty:
            return result
        closes = hist['Close']
        for w in windows:
            if len(closes) >= w:
                # Use pandas rolling window for standard SMA calculation
                result[f"MA_{w}"] = round(closes.rolling(window=w).mean().iloc[-1], 2)
        return result

    def calculate_ma_delta(self, current_price, avg):
        """Return the percentage delta: (current_price - avg) / avg."""
        if avg is None or current_price is None or avg == 0:
            return None
        delta = (current_price - avg) / avg
        return round(delta, 4)  # Return as float, e.g. 0.0521 for 5.21%

    @staticmethod
    def generate_yf_option_url(ticker, date_str):
        """Generate Yahoo Finance option chain URL with straddle view for a specific date."""
        if not date_str or not ticker:
            return None
        try:
            # Convert YYYY-MM-DD to unix timestamp at UTC midnight
            from datetime import timezone
            dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            timestamp = int(dt.timestamp())
            return f"https://finance.yahoo.com/quote/{ticker}/options/?date={timestamp}&straddle=true"
        except Exception:
            return None

    # ------------------------------------------------------------------
    @staticmethod
    def extract_abstract_terms(description, max_terms=1):
        """Return a short list of abstract merchant tokens from a description.

        Heuristics:
        - Prefer all-uppercase alpha tokens 2-12 chars long (e.g. 'BP', 'AMZN', 'WALMART').
        - Filter common stopwords.
        - Fallback to first alphanumeric token.
        """
        import re
        if not description:
            return []
        s = str(description).upper()
        # Common stopwords to ignore
        stopwords = {"THE", "AND", "OF", "IN", "AT", "ON", "FOR", "BY", "STORE", "LLC", "INC"}
        # Find tokens of 2-12 uppercase letters or digits (merchant codes often short or medium length)
        tokens = re.findall(r"\b[A-Z0-9&]{2,12}\b", s)
        # Keep order, remove stopwords and numeric-only tokens
        seen = set()
        out = []
        for t in tokens:
            if t in stopwords:
                continue
            if t in seen:
                continue
            # skip tokens that are all digits
            if t.isdigit():
                continue
            # Skip single-letter tokens
            if len(t) < 2:
                continue
            seen.add(t)
            out.append(t)
            if len(out) >= max_terms:
                break
        if out:
            return out
        # Fallback: take first alphanumeric word (up to 12 chars)
        m = re.search(r"[A-Z0-9&]{2,12}", s)
        if m:
            return [m.group(0)]
        words = re.findall(r"\w+", s)
        return [words[0][:12].upper()] if words else []

    @staticmethod
    def update_category_rules(json_path, description, category, max_terms=1):
        """Safely add minimal abstract term(s) from description to a category rules JSON.

        - `json_path` can be a Path or string.
        - Adds only short abstract tokens (from `extract_abstract_terms`) to avoid storing full verbose descriptions.
        - Does not overwrite existing mappings with a different category; logs and skips instead.
        - Ensures no duplicates.
        Returns the updated mapping dict.
        """
        import json
        json_path = Path(json_path)
        try:
            if json_path.exists():
                data = json.loads(json_path.read_text(encoding="utf-8"))
                if not isinstance(data, dict):
                    data = {}
            else:
                data = {}
        except Exception:
            data = {}

        terms = StockLiveComparison.extract_abstract_terms(description, max_terms=max_terms)
        changed = False
        for t in terms:
            if not t:
                continue
            if t in data:
                if data[t] != category:
                    logging.info(f"Existing mapping for {t} -> {data[t]} (requested {category}); skipping")
                # else mapping already correct
            else:
                data[t] = category
                changed = True

        if changed:
            try:
                json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
                logging.info(f"Updated category rules written to {json_path}")
            except Exception as e:
                logging.error(f"Failed to write category rules to {json_path}: {e}")
        return data

    # ------------------------------------------------------------------
    @staticmethod
    def get_latest_spreadsheet(directory: Path, base_name="AI_Stock_Live_Comparison_"):
        files = list(directory.glob(f"{base_name}*.xlsx"))
        if not files:
            return None, None
        files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        latest_file = files[0]
        file_time = datetime.fromtimestamp(latest_file.stat().st_mtime)
        return latest_file, file_time

    @staticmethod
    def is_viable_spreadsheet_file(path: Path, min_bytes=10 * 1024):
        """Return True when a report file is large enough to be used as a merge source."""
        try:
            return bool(path and path.exists() and path.stat().st_size >= int(min_bytes))
        except OSError:
            return False

    @classmethod
    def get_latest_viable_spreadsheet(cls, directory: Path, base_name="AI_Stock_Live_Comparison_", min_bytes=10 * 1024):
        """Return the latest report file that passes the minimum-size viability guard."""
        files = list(directory.glob(f"{base_name}*.xlsx"))
        if not files:
            return None, None
        files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        for candidate in files:
            if cls.is_viable_spreadsheet_file(candidate, min_bytes=min_bytes):
                return candidate, datetime.fromtimestamp(candidate.stat().st_mtime)
        return None, None

    @staticmethod
    def parse_report_date(report_path: Path, base_name="AI_Stock_Live_Comparison_"):
        match = re.match(rf"{re.escape(base_name)}(\d{{8}})_\d{{6}}\.xlsx$", report_path.name)
        if not match:
            return None
        try:
            return datetime.strptime(match.group(1), "%Y%m%d").date()
        except ValueError:
            return None

    @classmethod
    def get_latest_spreadsheet_for_date(cls, directory: Path, target_date, base_name="AI_Stock_Live_Comparison_"):
        files = list(directory.glob(f"{base_name}*.xlsx"))
        same_day = [f for f in files if cls.parse_report_date(f, base_name=base_name) == target_date]
        if not same_day:
            return None, None
        same_day.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        latest_file = same_day[0]
        return latest_file, datetime.fromtimestamp(latest_file.stat().st_mtime)

    # ------------------------------------------------------------------
    @staticmethod
    def closest_expiration(option_dates, target_days):
        target = datetime.now() + pd.Timedelta(days=target_days)
        dates = [datetime.strptime(d, "%Y-%m-%d") for d in option_dates]
        if not dates:
            return None
        return min(dates, key=lambda d: abs((d - target).days)).strftime("%Y-%m-%d")

    # ------------------------------------------------------------------
    def get_otm_call_yield(self, chain, current_price, target_days, otm_pct=6):
        exp_date = self.closest_expiration(chain.options, target_days)
        if not exp_date:
            return None, None, None
        calls = chain.option_chain(exp_date).calls
        target_strike = current_price * (1 + otm_pct / 100)
        calls = calls[calls['strike'] >= target_strike]
        if calls.empty:
            return None, None, None
        call = calls.iloc[0]
        yield_pct = (call['lastPrice'] / current_price) * 100
        return round(yield_pct, 2), call['strike'], exp_date

    # ------------------------------------------------------------------
    def get_otm_put_price(self, chain, current_price, target_days, otm_pct=6):
        exp_date = self.closest_expiration(chain.options, target_days)
        if not exp_date:
            return None, None
        puts = chain.option_chain(exp_date).puts
        target_strike = current_price * (1 - otm_pct / 100)
        puts = puts[puts['strike'] <= target_strike]
        if puts.empty:
            return None, None
        put = puts.iloc[-1]
        return put['lastPrice'], exp_date

    # ------------------------------------------------------------------
    def is_recent(self, row):
        try:
            last_update = pd.to_datetime(row.get("Last Update"))
            if pd.isnull(last_update):
                return False
            age = (self.now - last_update).total_seconds() / 3600
            return age < self.max_age_hours
        except Exception:
            return False

    # ------------------------------------------------------------------
    def fetch_ticker_record(self, ticker, info, hist, chain):
        ticker_hist = hist if hist is not None else None
        yoy = None
        prev_close = None
        if ticker_hist is not None and not ticker_hist.empty:
            yoy = (ticker_hist['Close'].iloc[-1] - ticker_hist['Close'].iloc[0]) / ticker_hist['Close'].iloc[0] * 100
            prev_close = ticker_hist['Close'].iloc[-2]

        current_price = info.get("regularMarketPrice") or info.get("currentPrice")
        day_change = None
        if current_price and prev_close:
            day_change = (current_price - prev_close) / prev_close * 100

        call3, _, call_date3 = self.get_otm_call_yield(chain, current_price, 90)
        call6, strike6, call_date6 = self.get_otm_call_yield(chain, current_price, 180)
        call12, _, call_date12 = self.get_otm_call_yield(chain, current_price, 365)
        put_price, put_date12 = self.get_otm_put_price(chain, current_price, 365)
        analyst_target = info.get("targetMeanPrice")

        ex_div_raw = info.get("exDividendDate")
        if ex_div_raw:
            try:
                ex_div_date = datetime.fromtimestamp(ex_div_raw).strftime("%Y-%m-%d")
            except Exception:
                ex_div_date = str(ex_div_raw)
        else:
            ex_div_date = None

        annual_yield_put = None
        if put_price and current_price:
            annual_yield_put = round((put_price / current_price) * 100, 2)

        annual_yield_call = None
        if call12 and current_price:
            annual_yield_call = round((call12 / current_price) * 100, 2)

        # New Indicators: EMA, HMA, TSMOM, RSI, ATR
        ema_20 = self.calculate_ema(ticker_hist['Close'], span=20) if ticker_hist is not None else None
        hma_20 = self.calculate_hma(ticker_hist['Close'], window=20) if ticker_hist is not None else None
        # TSMOM 60-day lookback (Updated from 45)
        tsmom_60 = self.calculate_tsmom(ticker_hist['Close'], lookback=60) if ticker_hist is not None else None
        
        rsi_14 = self.calculate_rsi(ticker_hist['Close'], period=14) if ticker_hist is not None else None
        atr_14 = self.calculate_atr(ticker_hist['High'], ticker_hist['Low'], ticker_hist['Close'], period=14) if ticker_hist is not None else None

        # Price Action Analysis
        price_action = {}
        if ticker_hist is not None and not ticker_hist.empty:
            try:
                price_action = PriceActionService.analyze_ticker(ticker_hist)
            except Exception as e:
                self.logger.error(f"Price Action Error for {ticker}: {e}")

        # Moving averages and highlight status
        ma_windows = (30, 60, 120, 200)
        ma_dict = self.calculate_moving_averages(ticker_hist, windows=ma_windows)
        highlight_dict = {f"MA_{w}_highlight": self.calculate_ma_delta(current_price, ma_dict[f"MA_{w}"])
                  for w in ma_windows}
        
        # Calculate highlights for new indicators
        ema_20_highlight = self.calculate_ma_delta(current_price, ema_20)
        # HMA Highlight: Price - HMA (diff)
        hma_20_highlight = current_price - hma_20 if hma_20 is not None and current_price is not None else None
        # TSMOM Highlight: Just the value itself 
        tsmom_60_highlight = tsmom_60

        record = {
            "Ticker": ticker,
            "Company Name": info.get("longName") or info.get("shortName") or ticker,
            "Current Price": current_price,
            "1D % Change": f"{day_change:.2f}%" if day_change is not None else None,
            "Market Cap (T$)": info.get("marketCap") / 1e12 if info.get("marketCap") else None,
            "P/E": info.get("trailingPE"),
            "YoY Price %": f"{yoy:.1f}%" if yoy is not None else None,
            "EMA_20": ema_20,
            "HMA_20": hma_20,
            "TSMOM_60": tsmom_60,
            "RSI_14": rsi_14,
            "ATR_14": atr_14,
            "EMA_20_highlight": ema_20_highlight,
            "HMA_20_highlight": hma_20_highlight,
            "TSMOM_60_highlight": tsmom_60_highlight,
            "Ex-Div Date": ex_div_date,
            "Div Yield": info.get("dividendYield"),
            "Analyst 1-yr Target": analyst_target,
            "1-yr 6% OTM PUT Price": put_price,
            "Annual Yield Put Prem": annual_yield_put,
            "3-mo Call Yield": call3,
            "6-mo Call Yield": call6,
            "1-yr Call Yield": call12,
            "Annual Yield Call Prem": annual_yield_call,
            "6-mo Call Strike": strike6,
            "Error": None,
            "Last Update": self.now.strftime("%Y-%m-%d %H:%M:%S"),
            "_PutExpDate_365": put_date12,
            "_CallExpDate_365": call_date12,
            "_CallExpDate_90": call_date3,
            "_CallExpDate_180": call_date6,
            "Price Action": price_action,
        }
        # Add moving averages and highlight info
        record.update(ma_dict)
        record.update(highlight_dict)

        # Company profile + recent news (stored as sub-document)
        news_items = []
        raw_news = []
        if self.fetch_profile_news:
            try:
                raw_news = chain.news or []
            except Exception as e:
                self.logger.warning(f"stock_live_comparison.fetch_ticker_record - Could not fetch news for {ticker}: {e}")
                raw_news = []
        else:
            # Preserve preloaded/mocked news without triggering live HTTP profile-news calls.
            chain_dict = getattr(chain, "__dict__", {}) if chain is not None else {}
            maybe_news = chain_dict.get("news")
            if isinstance(maybe_news, list):
                raw_news = maybe_news

        if isinstance(raw_news, list) and raw_news:
            news_items = [
                {
                    "title": n.get("title", ""),
                    "publisher": n.get("publisher", ""),
                    "link": n.get("link", ""),
                    "published_at": datetime.fromtimestamp(n["providerPublishTime"]).strftime("%Y-%m-%d %H:%M")
                    if n.get("providerPublishTime") else "",
                }
                for n in raw_news[:5]
                if isinstance(n, dict)
            ]

        record["profile"] = {
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "description": info.get("longBusinessSummary", ""),
            "style": info.get("quoteType", ""),
            "category": info.get("category", ""),
            "exchange": info.get("exchange", ""),
            "country": info.get("country", ""),
            "employees": info.get("fullTimeEmployees"),
            "website": info.get("website", ""),
            "recommendation": info.get("recommendationKey", ""),
            "analyst_opinions": info.get("numberOfAnalystOpinions"),
            "beta": info.get("beta"),
            "forward_pe": info.get("forwardPE"),
            "price_to_book": info.get("priceToBook"),
            "roe": info.get("returnOnEquity"),
            "debt_to_equity": info.get("debtToEquity"),
            "earnings_growth": info.get("earningsGrowth"),
            "revenue_growth": info.get("revenueGrowth"),
            "news": news_items,
        }

        return record

    @staticmethod
    def calculate_ema(series, span=20):
        """Calculate Exponential Moving Average."""
        if series is None or series.empty or len(series) < span:
            return None
        return round(series.ewm(span=span, adjust=False).mean().iloc[-1], 2)

    @staticmethod
    def weighted_moving_average(series, window):
        """Calculate Weighted Moving Average (helper for HMA)."""
        import numpy as np
        weights = np.arange(1, window + 1)
        return series.rolling(window).apply(
            lambda x: np.dot(x, weights) / weights.sum(), 
            raw=True
        )

    def calculate_hma(self, series, window=20):
        """Calculate Hull Moving Average."""
        import numpy as np
        if series is None or series.empty or len(series) < window:
            return None
        try:
            # Step 1: WMA(n/2)
            wma_half = self.weighted_moving_average(series, window // 2)
            # Step 2: WMA(n)
            wma_full = self.weighted_moving_average(series, window)
            # Step 3: 2 * WMA(n/2) - WMA(n)
            raw_hma = (2 * wma_half) - wma_full
            # Step 4: WMA(sqrt(n))
            lag = int(np.sqrt(window))
            hma = self.weighted_moving_average(raw_hma, lag)
            return round(hma.iloc[-1], 2)
        except Exception as e:
            return None

    @staticmethod
    def calculate_tsmom(series, lookback=60):
        """Calculate Time Series Momentum (Return) over lookback period (default 60 days)."""
        if series is None or series.empty or len(series) <= lookback:
            return None
        try:
            # (Price_t / Price_{t-lookback}) - 1
            current = series.iloc[-1]
            past = series.iloc[-(lookback + 1)] # ensure we use t-lookback
            if past == 0:
                return None
            return round((current / past) - 1, 4)
        except Exception:
            return None

    @staticmethod
    def calculate_rsi(series, period=14):
        """Calculate Relative Strength Index (RSI)."""
        if series is None or series.empty or len(series) < period:
            return None
        try:
            delta = series.diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()

            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            return round(rsi.iloc[-1], 2)
        except Exception:
            return None

    @staticmethod
    def calculate_atr(high, low, close, period=14):
        """Calculate Average True Range (ATR)."""
        if high is None or low is None or close is None or len(close) < period:
            return None
        try:
            import numpy as np
            # TR = Max(High-Low, Abs(High-Close_prev), Abs(Low-Close_prev))
            # Pandas approach
            
            # Need strict alignment
            h = high[-period*2:] 
            l = low[-period*2:]
            c = close[-period*2:]
            if len(c) < 2: return None
            
            c_prev = c.shift(1)
            tr1 = h - l
            tr2 = (h - c_prev).abs()
            tr3 = (l - c_prev).abs()
            
            tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
            atr = tr.rolling(window=period).mean().iloc[-1]
            return round(atr, 2)
        except Exception:
            return None

    # ------------------------------------------------------------------
    def fetch_data(self, tickers_to_fetch):
        tickers_to_fetch = StockLiveComparison.unique_tickers(tickers_to_fetch)
        if not tickers_to_fetch:
            return []
        total = len(tickers_to_fetch)
        logging.info(f"Downloading historical data for {total} tickers...")
        hist = self.download_history_batched(tickers_to_fetch)
        time.sleep(1)
        try:
            tickers_obj = yf.Tickers(" ".join(tickers_to_fetch))
        except Exception as e:
            return [
                {
                    "Ticker": t,
                    "Error": str(e),
                    "Last Update": self.now.strftime("%Y-%m-%d %H:%M:%S"),
                }
                for t in tickers_to_fetch
            ]
        records = []
        max_attempts = 4
        for idx, t in enumerate(tickers_to_fetch, start=1):
            attempt = 0
            while attempt < max_attempts:
                attempt += 1
                try:
                    if attempt == 1 and (idx == 1 or idx % 10 == 0 or idx == total):
                        logging.info("Stock analysis progress: %s/%s (ticker=%s)", idx, total, t)
                    logging.debug(f"Processing ticker {t} (Attempt {attempt}/{max_attempts})...")
                    # Visual interaction for console (optional, keeps user happy)
                    # print(f"Processing {t}...", end="\r", flush=True) 
                    # Actually logging.info might be too noisy if list is long, but user asked for visuals.
                    # Best to stick to requested debug log for now, relying on YF progress bar for download.
                    
                    self.throttle_yf_requests()
                    time.sleep(1 + max(0, attempt - 1))
                    ticker_obj = tickers_obj.tickers.get(t)
                    if ticker_obj is None:
                        raise KeyError(f"Ticker '{t}' missing from yfinance.Tickers result")
                    info = ticker_obj.info
                    ticker_hist = hist.get(t)
                    chain = ticker_obj
                    record = self.fetch_ticker_record(t, info, ticker_hist, chain)
                    records.append(record)
                    self._consecutive_429 = 0
                    break
                except Exception as e:
                    retryable = self.is_retryable_yf_error(e)
                    if self.is_http_429_error(e):
                        self._consecutive_429 += 1
                        self.apply_rate_limit_cooldown(self._consecutive_429)
                    should_retry = retryable and attempt < max_attempts
                    if should_retry:
                        backoff = min(90, 3 * (2 ** (attempt - 1))) + random.uniform(0.1, 0.9)
                        logging.warning(
                            "Stock analysis ticker retry: %s attempt=%s/%s backoff_sec=%s error=%s",
                            t,
                            attempt,
                            max_attempts,
                            round(backoff, 2),
                            e,
                        )
                        time.sleep(backoff)
                        continue

                    logging.warning(
                        "Stock analysis ticker failed: %s attempt=%s/%s retryable=%s error=%s",
                        t,
                        attempt,
                        max_attempts,
                        retryable,
                        e,
                    )
                    record = {
                        "Ticker": t,
                        "Error": str(e),
                        "Last Update": self.now.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                    records.append(record)
                    break
        logging.info("Stock analysis fetch complete: %s records produced for %s requested tickers.", len(records), total)
        return records

    def download_history_batched(self, tickers_to_fetch):
        """Download history in small batches, with local cache and paced pauses."""
        hist_by_ticker = {}
        to_download = []
        for ticker in tickers_to_fetch:
            cached = self.load_cached_history(ticker)
            if cached is not None:
                hist_by_ticker[ticker] = cached
            else:
                to_download.append(ticker)

        if not to_download:
            return hist_by_ticker

        batches = [
            to_download[i : i + self.download_batch_size]
            for i in range(0, len(to_download), self.download_batch_size)
        ]
        total_batches = len(batches)

        for idx, batch in enumerate(batches, start=1):
            self.throttle_yf_requests()
            logging.info(
                "Stock analysis history batch %s/%s size=%s",
                idx,
                total_batches,
                len(batch),
            )
            try:
                batch_hist = yf.download(
                    batch,
                    period="1y",
                    group_by="ticker",
                    threads=True,
                    auto_adjust=False,  # explicit to avoid FutureWarning
                    progress=False,
                )
            except Exception as e:
                logging.warning(
                    "Stock analysis history batch failed batch=%s/%s error=%s",
                    idx,
                    total_batches,
                    e,
                )
                batch_hist = None

            for ticker in batch:
                ticker_hist = self.extract_ticker_history(batch_hist, ticker, batch_len=len(batch))
                hist_by_ticker[ticker] = ticker_hist
                if ticker_hist is not None and not ticker_hist.empty:
                    self.write_cached_history(ticker, ticker_hist)

            if idx < total_batches and self.batch_pause_sec > 0:
                time.sleep(self.batch_pause_sec)

        return hist_by_ticker

    def history_cache_path(self, ticker):
        safe = re.sub(r"[^A-Za-z0-9_.-]", "_", str(ticker))
        return self.history_cache_dir / f"{safe}.csv"

    def is_cache_fresh(self, cache_path):
        if not cache_path.exists():
            return False
        age_hours = (time.time() - cache_path.stat().st_mtime) / 3600.0
        return age_hours <= self.history_cache_ttl_hours

    def load_cached_history(self, ticker):
        cache_path = self.history_cache_path(ticker)
        if not self.is_cache_fresh(cache_path):
            return None
        try:
            df = pd.read_csv(cache_path, index_col=0, parse_dates=True)
            # We expect OHLCV columns; treat malformed cache as miss.
            required = {"Open", "High", "Low", "Close"}
            if not required.issubset(set(df.columns)):
                return None
            return df
        except Exception:
            return None

    def write_cached_history(self, ticker, df):
        cache_path = self.history_cache_path(ticker)
        try:
            df.to_csv(cache_path)
        except Exception as e:
            logging.warning("Stock analysis cache write failed ticker=%s error=%s", ticker, e)

    @staticmethod
    def extract_ticker_history(batch_hist, ticker, batch_len):
        if batch_hist is None:
            return None
        try:
            # Single-ticker batch returns a regular OHLC DataFrame.
            if batch_len == 1 and isinstance(batch_hist, pd.DataFrame):
                return batch_hist
            # Multi-ticker batch usually returns MultiIndex columns keyed by ticker.
            if isinstance(batch_hist, pd.DataFrame) and isinstance(batch_hist.columns, pd.MultiIndex):
                return batch_hist[ticker] if ticker in batch_hist.columns.get_level_values(0) else None
            # Fallback for dict-like stubs in tests.
            if isinstance(batch_hist, dict):
                return batch_hist.get(ticker)
        except Exception:
            return None
        return None

    @staticmethod
    def is_retryable_yf_error(exc):
        """Best-effort classifier for transient yfinance/network errors."""
        msg = str(exc).lower()
        if "429" in msg or "too many requests" in msg or "rate limit" in msg:
            return True
        transient_markers = (
            "temporarily unavailable",
            "timed out",
            "timeout",
            "connection reset",
            "connection aborted",
            "service unavailable",
        )
        return any(marker in msg for marker in transient_markers)

    @staticmethod
    def is_http_429_error(exc):
        msg = str(exc).lower()
        return "429" in msg or "too many requests" in msg or "rate limit" in msg

    def throttle_yf_requests(self):
        """Throttle outbound yfinance calls to avoid burst behavior."""
        now_ts = time.time()
        if now_ts < self._next_request_not_before_ts:
            time.sleep(self._next_request_not_before_ts - now_ts)
            now_ts = time.time()
        self._next_request_not_before_ts = now_ts + max(0.0, self.min_request_interval_sec)

    def apply_rate_limit_cooldown(self, consecutive_hits):
        """Apply a global cooldown window after 429 responses."""
        cooldown = min(300, 20 * (2 ** max(0, consecutive_hits - 1)))
        cooldown += random.uniform(0.5, 2.0)
        self._next_request_not_before_ts = max(self._next_request_not_before_ts, time.time() + cooldown)
        logging.warning(
            "Stock analysis 429 cooldown engaged: consecutive=%s cooldown_sec=%s",
            consecutive_hits,
            round(cooldown, 2),
        )

    # ------------------------------------------------------------------
    def merge_with_existing(self, df_existing, tickers_to_fetch):
        df_new = pd.DataFrame(self.records)
        df_existing = df_existing[~df_existing['Ticker'].isin(tickers_to_fetch)]
        return pd.concat([df_existing, df_new], ignore_index=True)

    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    def get_missing_or_outdated_tickers(self, df_existing):
        """Return tickers missing form the DataFrame, older than max_age_hours, or having missing MA data."""
        missing = []
        ma_cols = ["MA_30", "MA_60", "MA_120", "MA_200"]
        # Include new indicators in validity check
        new_cols = ["EMA_20", "HMA_20", "TSMOM_60", "RSI_14", "ATR_14"]
        all_required_cols = ma_cols + new_cols
        
        for t in self.tickers:
            row = df_existing[df_existing["Ticker"] == t]
            if row.empty:
                missing.append(t)
                continue
            try:
                # Check age
                last = pd.to_datetime(row.iloc[0]["Last Update"])
                if pd.isna(last) or (
                    (self.now - last).total_seconds() / 3600 >= self.max_age_hours
                ):
                    missing.append(t)
                    continue
                
                # Check for missing Moving Averages and New Indicators (NaN)
                # If any required column is missing or NaN, consider it outdated to force re-fetch
                for col in all_required_cols:
                    if col not in row.columns or pd.isna(row.iloc[0][col]):
                        missing.append(t)
                        break
                else:
                    # Check for missing Expiration Date fields (Validation for YF Links)
                    # Use _PutExpDate_365 as a proxy for all new date fields
                    if "_PutExpDate_365" not in row.columns or pd.isna(row.iloc[0]["_PutExpDate_365"]):
                         missing.append(t)
                    else:
                        # Check for legacy "green"/"red" strings in highlight columns
                        # Only check if the previous check didn't already mark it as missing
                        for ma in ma_cols:
                            hl_col = f"{ma}_highlight"
                            if hl_col in row.columns:
                                val = row.iloc[0][hl_col]
                                if isinstance(val, str) and val in ["green", "red"]:
                                    missing.append(t)
                                    break
                        
            except Exception:
                missing.append(t)
        return missing

    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    def add_ratio_column(self, df):
        """Add Call/Put Skew column to DataFrame using vectorized operations."""
        ratio_col_name = "Call/Put Skew"
        try:
            # Ensure columns exist
            for col in ["Annual Yield Put Prem", "Annual Yield Call Prem"]:
                if col not in df.columns:
                    df[col] = None

            # Determine insertion index (after Call Prem)
            try:
                call_col = df.columns.get_loc("Annual Yield Call Prem") + 1
                put_col = df.columns.get_loc("Annual Yield Put Prem") + 1
            except Exception:
                # If columns are missing or index fails, just append to end
                call_col = len(df.columns)
                put_col = len(df.columns)

            # Convert to numeric, forcing non-numeric to NaN
            s_call = pd.to_numeric(df["Annual Yield Call Prem"], errors='coerce')
            s_put = pd.to_numeric(df["Annual Yield Put Prem"], errors='coerce')

            # Calculate Skew (Vectorized)
            # handle division by zero (result becomes inf, then replace with NaN)
            skew_series = s_call / s_put
            skew_series = skew_series.replace([float('inf'), float('-inf')], float('nan'))
            
            # Insert or Update Column
            if ratio_col_name in df.columns:
                df[ratio_col_name] = skew_series
            else:
                # Insert at specific position if safe, else assign
                if call_col < len(df.columns):
                    df.insert(call_col, ratio_col_name, skew_series)
                else:
                    df[ratio_col_name] = skew_series

            return df, put_col, call_col

        except Exception as e:
            logging.error(f"Error in add_ratio_column: {e}")
            return df, None, None

    # ------------------------------------------------------------------
    def upsert_ratio_column(self, df, put_col_name="Annual Yield Put Prem", call_col_name="Annual Yield Call Prem", ratio_col_name="Call/Put Skew"):
        """Upsert the Call/Put Skew column, print errors, and continue."""
        try:
            if ratio_col_name in df.columns:
                logging.debug(f'Column "{ratio_col_name}" already exists. Updating values.')
            else:
                # Insert after call_col_name
                if call_col_name in df.columns:
                    call_col = df.columns.get_loc(call_col_name) + 1
                    df.insert(call_col, ratio_col_name, None)
                    logging.debug(f'Column "{ratio_col_name}" inserted.')
                else:
                    df[ratio_col_name] = None
                    logging.debug(f'Column "{ratio_col_name}" added at end (call column not found).')

            # Update values
            df[ratio_col_name] = df.apply(
                lambda row: (
                    row[call_col_name] / row[put_col_name]
                    if row[put_col_name] not in [0, None, ""] and row[call_col_name] not in [None, ""] else None
                ),
                axis=1
            )
            return df
        except Exception as e:
            logging.error(f"Error in upsert_ratio_column: {e}")
            return df

    # ------------------------------------------------------------------
    def save_to_excel(self, df, put_col, call_col):
        # Ensure all MA and new columns exist in DataFrame before saving
        ma_windows = [30, 60, 120, 200]
        new_cols = ["EMA_20", "HMA_20", "TSMOM_60", "RSI_14", "ATR_14"]
        all_cols = new_cols + [f"MA_{w}" for w in ma_windows]
        
        for col in all_cols:
            if col not in df.columns:
                df[col] = None

        # Reorder columns: Place new columns before MA_30
        initial_cols = [c for c in df.columns if c not in all_cols]
        # Find insertion point after "YoY Price %"
        ordered_cols = []
        if "YoY Price %" in initial_cols:
             idx = initial_cols.index("YoY Price %") + 1
             ordered_cols = initial_cols[:idx] + new_cols + [f"MA_{w}" for w in ma_windows] + initial_cols[idx:]
        else:
             ordered_cols = initial_cols + new_cols + [f"MA_{w}" for w in ma_windows]
             
        # Filter and fill
        ordered_cols = [c for c in ordered_cols if c in df.columns]
        remaining = [c for c in df.columns if c not in ordered_cols]
        ordered_cols.extend(remaining)
        
        df = df[ordered_cols]

        df = self.sort_dataframe_for_excel(df)
        df.to_excel(self.filename, index=False)
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws.row_dimensions[1].height = None

        if "Ticker" in df.columns:
            ticker_col_idx = df.columns.get_loc("Ticker") + 1
        else:
            ticker_col_idx = None

        # Re-calc indices after reorder
        put_col = df.columns.get_loc("Annual Yield Put Prem") + 1 if "Annual Yield Put Prem" in df.columns else None
        call_col = df.columns.get_loc("Annual Yield Call Prem") + 1 if "Annual Yield Call Prem" in df.columns else None
        ratio_col = df.columns.get_loc("Call/Put Skew") + 1 if "Call/Put Skew" in df.columns else None

        # Hyperlinks ONLY (Removed IFERROR formula overwrite)
        for i in range(2, ws.max_row + 1):
            
            # Add Google Finance Hyperlink to Ticker Column
            if ticker_col_idx:
                ticker_cell = ws.cell(row=i, column=ticker_col_idx)
                ticker_val = ticker_cell.value
                if ticker_val:
                    url = f"https://www.google.com/finance?q={str(ticker_val)}"
                    ticker_cell.hyperlink = url
                    ticker_cell.style = "Hyperlink"

            # Add Yahoo Finance Option Chain Hyperlinks to Yield Columns
            # Define mapping of visible column name -> record key for expiration date
            link_map = {
                "1-yr 6% OTM PUT Price": "_PutExpDate_365",
                "1-yr Call Yield": "_CallExpDate_365",
                "3-mo Call Yield": "_CallExpDate_90",
                "6-mo Call Yield": "_CallExpDate_180",
                "6-mo Call Strike": "_CallExpDate_180",
            }
            
            # Pre-calculate column indices to avoid repeated lookups
            col_indices = {}
            for col_name, date_key in link_map.items():
                if col_name in df.columns:
                     # 1-based index for openpyxl corresponding to DF column
                     col_indices[df.columns.get_loc(col_name) + 1] = date_key

            # Apply links row by row
            if col_indices:
                for col_idx, date_key in col_indices.items():
                    # Row i in Excel corresponds to df index i-2
                    if 0 <= i-2 < len(df):
                        row_data = df.iloc[i-2]
                        exp_date = row_data.get(date_key)
                        ticker = row_data.get("Ticker")
                        
                        if exp_date and ticker:
                            yf_url = self.generate_yf_option_url(ticker, exp_date)
                            if yf_url:
                                cell = ws.cell(row=i, column=col_idx)
                                if cell.value:
                                    cell.hyperlink = yf_url
                                    cell.style = "Hyperlink"

        # Call/Put Skew Conditional Formatting
        if ratio_col:
            for i in range(2, ws.max_row + 1):
                cell = ws.cell(row=i, column=ratio_col)
                # Check value from DF (since Excel formula isn't evaluated by openpyxl reading)
                if 0 <= i-2 < len(df):
                    row_data = df.iloc[i-2]
                    val = row_data.get("Call/Put Skew")
                    # Should be a number.
                    if val is not None and isinstance(val, (int, float)):
                        if val > 1.2:
                            cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                        elif val < 0.75:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red

        # SMA Conditional Formatting
        for w in ma_windows:
            col_name = f"MA_{w}"
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for i in range(2, ws.max_row + 1):
                    avg_cell = ws.cell(row=i, column=col_idx)
                    price_col_idx = df.columns.get_loc("Current Price") + 1 if "Current Price" in df.columns else None
                    if price_col_idx:
                        price_cell = ws.cell(row=i, column=price_col_idx)
                        try:
                            avg = avg_cell.value
                            price = price_cell.value
                            if avg is not None and price is not None:
                                if price <= avg * (1 - self.highlight_threshold):
                                    avg_cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
                                elif price >= avg * (1 + self.highlight_threshold):
                                    avg_cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
                        except Exception:
                            pass
            
            # Format highlight column as percentage
            hl_col_name = f"MA_{w}_highlight"
            if hl_col_name in df.columns:
                hl_col_idx = df.columns.get_loc(hl_col_name) + 1
                for i in range(2, ws.max_row + 1):
                    cell = ws.cell(row=i, column=hl_col_idx)
                    cell.number_format = '0.0%'

        # EMA Color Coding (20 day)
        # Price > 0.5% color green, < -0.5% RED
        if "EMA_20" in df.columns:
            ema_col_idx = df.columns.get_loc("EMA_20") + 1
            hl_col_name = "EMA_20_highlight"
            
            for i in range(2, ws.max_row + 1):
                cell = ws.cell(row=i, column=ema_col_idx)
                if 0 <= i-2 < len(df) and hl_col_name in df.columns:
                     val = df.iloc[i-2][hl_col_name]
                     if val is not None:
                         if val > 0.005: # > 0.5%
                             cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                         elif val < -0.005: # < -0.5%
                             cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red

        # HMA Color Coding (20 day)
        # Price > HMA Green, Price < HMA Red
        if "HMA_20" in df.columns:
            hma_col_idx = df.columns.get_loc("HMA_20") + 1
            hl_col_name = "HMA_20_highlight" # Price - HMA
            
            for i in range(2, ws.max_row + 1):
                cell = ws.cell(row=i, column=hma_col_idx)
                if 0 <= i-2 < len(df) and hl_col_name in df.columns:
                     val = df.iloc[i-2][hl_col_name]
                     if val is not None:
                         if val > 0:
                             cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                         elif val < 0:
                             cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
        
        # TSMOM Color Coding (60 day)
        # Green > 2%, Red < -2%
        if "TSMOM_60" in df.columns:
             tsmom_col_idx = df.columns.get_loc("TSMOM_60") + 1
             # Format as percentage
             for i in range(2, ws.max_row + 1):
                cell = ws.cell(row=i, column=tsmom_col_idx)
                cell.number_format = '0.0%'
                if 0 <= i-2 < len(df):
                    val = df.iloc[i-2]["TSMOM_60"] # TSMOM_60_highlight is same as TSMOM_60
                    if val is not None:
                        if val > 0.02: # > 2%
                            cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                        elif val < -0.02: # < -2%
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A2"  # Freeze the first row

        wb.save(self.filename)

    # ------------------------------------------------------------------
    def sort_dataframe_for_excel(self, df):
        """Sort the DataFrame by Last Update (descending) and Ticker (ascending)."""
        try:
            if "Last Update" in df.columns and "Ticker" in df.columns:
                df = df.sort_values(by=["Last Update", "Ticker"], ascending=[False, True])
                df = df.sort_values(by=["Last Update", "Ticker"], ascending=[False, True])
            else:
                logging.warning("Warning: 'Last Update' or 'Ticker' column missing. Skipping sort.")
            return df
        except Exception as e:
            logging.error(f"Error in sort_dataframe_for_excel: {e}")
            return df

    # ------------------------------------------------------------------
    def upsert_to_mongo(self, df):
        """
        Upsert each row of the DataFrame into the canonical `stock_data` collection
        using `Ticker` as the unique key used by ticker detail APIs.

        This method preserves required detail payload fields (for modal reliability)
        when a row originated from spreadsheet merge data that omits nested fields
        such as `profile`.
        """
        try:
            db = AiStockDatabase(collection_name="stock_data")
            price_history_db = AiStockDatabase(collection_name="instrument_price_history")
            self.ensure_price_history_indexes(price_history_db.collection)
            records = df.to_dict(orient="records")
            required = self.required_detail_fields()
            now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for record in records:
                try:
                    ticker = str(record.get("Ticker") or "").strip().upper()
                    if not ticker:
                        continue

                    # Spreadsheet merge rows can omit nested detail fields; preserve
                    # existing canonical values when incoming row is sparse.
                    existing = db.collection.find_one({"Ticker": ticker}) or {}
                    merged = self.merge_detail_record(record, existing)
                    merged["Ticker"] = ticker
                    merged["_last_persisted_at"] = now_ts
                    db.upsert_stock_record(merged, key_fields=("Ticker",))
                    history_record = self.build_price_history_record(merged, default_ts=now_ts)
                    if history_record:
                        price_history_db.upsert_stock_record(
                            history_record,
                            key_fields=("instrument_key", "timestamp", "source"),
                        )

                    missing = self.missing_required_detail_fields(merged, required_fields=required)
                    if missing:
                        logging.warning(
                            "stock_live_comparison.upsert_to_mongo - ticker=%s missing required detail fields: %s",
                            ticker,
                            ",".join(missing),
                        )
                except Exception as e:
                    logging.error(f"Error upserting record for {record.get('Ticker')}: {e}")
            logging.info(f"Upserted {len(records)} records to MongoDB.")
        except Exception as e:
            logging.error(f"Error connecting to MongoDB: {e}")

    @classmethod
    def ensure_price_history_indexes(cls, collection):
        if cls._price_history_indexes_ensured:
            return
        try:
            collection.create_index([("instrument_key", 1), ("timestamp", -1)])
            collection.create_index([("source", 1), ("timestamp", -1)])
            cls._price_history_indexes_ensured = True
        except Exception as exc:
            logging.warning("Unable to ensure instrument_price_history indexes: %s", exc)

    @staticmethod
    def build_price_history_record(record, default_ts=None):
        ticker = str((record or {}).get("Ticker") or "").strip().upper()
        if not ticker:
            return None
        timestamp = (record or {}).get("Last Update") or (record or {}).get("_last_persisted_at") or default_ts
        if not timestamp:
            return None
        return {
            "instrument_key": ticker,
            "instrument_type": "STK",
            "timestamp": str(timestamp),
            "price": (record or {}).get("Current Price"),
            "day_change_pct": (record or {}).get("1D % Change"),
            "source": "stock_live_comparison",
            "_ingested_at": default_ts or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def required_detail_fields():
        # Core fields used by ticker detail/header/modal tabs.
        return [
            "Ticker",
            "Company Name",
            "Current Price",
            "1D % Change",
            "Last Update",
            "Call/Put Skew",
            "Price Action",
            "profile",
        ]

    @staticmethod
    def merge_detail_record(incoming, existing):
        merged = dict(existing or {})
        merged.update(incoming or {})

        incoming_profile = (incoming or {}).get("profile")
        if (incoming_profile is None or incoming_profile == {}) and (existing or {}).get("profile"):
            merged["profile"] = (existing or {}).get("profile")

        incoming_price_action = (incoming or {}).get("Price Action")
        if (incoming_price_action is None or incoming_price_action == {}) and (existing or {}).get("Price Action"):
            merged["Price Action"] = (existing or {}).get("Price Action")

        return merged

    @staticmethod
    def missing_required_detail_fields(record, required_fields=None):
        required = required_fields or StockLiveComparison.required_detail_fields()
        missing = []
        for field in required:
            value = record.get(field)
            if value is None or value == "":
                missing.append(field)
            elif field == "profile" and isinstance(value, dict):
                # Require deterministic news key for profile-driven modal rendering.
                if "news" not in value:
                    missing.append("profile.news")
        return missing

    # ------------------------------------------------------------------
    @staticmethod
    def unique_tickers(tickers):
        """Return a list of unique tickers, preserving order."""
        # variable tickers_to_fetch in your code is a native Python list, not a pandas Series.
        seen = set()
        unique = []
        for t in tickers:
            if t not in seen:
                unique.append(t)
                seen.add(t)
        return unique

    @staticmethod
    def get_default_tickers():
        """Return the default list of tickers to track, sourcing from MongoDB if available."""
        default_list = sorted(list({
            "^IXIC", "^SPX", "SPXS", "^DJI",
            "AA", "AAPL", "AMAT", "AMD", "AMZN", "AVGO", "BHP", "BMY", "CCJ", "CEG", "COPP",
            "CPRX", "CRWD", "CRWV", "CVS", "CVX", "D", "DUK", "ENB", "ERO", "ETN",
            "F", "FDX", "FMNB", "GD", "GE", "GEV", "GOOG", "GOOGL", "FCX", 
            "IBM", "INTC", "IONQ", "JNJ", "JPM", "KMB", "KO", "LAC", "LRCX", "MCD", "META",
            "MO", "MRVL", "MSFT", "MU", "NEE", "NNE", "NUE", "NVDA",
            "OKE", "OLN", "ORCL", "PAAS", "PFE", "PLTR", "RIO", "SLB", "SMG", "SMR", "STLD",
            "SCCO", "T", "TECH", "TEM", "TMUS", "TSLA", "TSM", "UPS", 
            "V", "VZ", "VLO", "VSAT", "VST", "WM", "WMT", "XOM"
        }))

        try:
            db = AiStockDatabase()
            # Access system_config directly via the underlying client/db
            if hasattr(db, 'db'):
                config_col = db.db["system_config"]
                doc = config_col.find_one({"_id": "tracked_tickers"})
                
                if doc and "tickers" in doc:
                    logging.info(f"Loaded {len(doc['tickers'])} tickers from MongoDB.")
                    return sorted(doc["tickers"])
                else:
                    # Migration: Save default list to MongoDB
                    logging.info("No tracked tickers found in DB. Migrating defaults...")
                    config_col.update_one(
                        {"_id": "tracked_tickers"},
                        {"$set": {"tickers": default_list}},
                        upsert=True
                    )
                    return default_list
        except Exception as e:
            logging.error(f"Failed to load tickers from MongoDB, using defaults: {e}")
            
        return default_list

    def select_output_report_file(self, force_new_file=False, allow_create_if_missing=True):
        """Pick output report file based on trigger semantics.

        - force_new_file=True: always create a new timestamped file.
        - force_new_file=False:
            - reuse today's existing report if present
            - else reuse latest existing report when allow_create_if_missing=False
            - else create new timestamped report
        """
        if force_new_file:
            return self.output_dir / f"AI_Stock_Live_Comparison_{self.now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        today_file, _ = self.get_latest_spreadsheet_for_date(self.output_dir, self.now.date())
        if today_file:
            return today_file

        if not allow_create_if_missing:
            latest_file, _ = self.get_latest_spreadsheet(self.output_dir)
            if latest_file:
                return latest_file

        return self.output_dir / f"AI_Stock_Live_Comparison_{self.now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    def is_suspicious_record_count(self, record_count):
        """Detect suspiciously small report outputs to prevent accidental overwrite churn."""
        expected = len(self.tickers)
        if expected < 20:
            return False
        minimum_allowed = max(5, int(expected * 0.2))
        return int(record_count) < minimum_allowed

    def run(self, force_new_file=False, allow_create_if_missing=True):
        self.now = datetime.now()
        self.filename = self.select_output_report_file(
            force_new_file=force_new_file,
            allow_create_if_missing=allow_create_if_missing,
        )
        self.latest_file, _ = self.get_latest_spreadsheet(self.output_dir)
        self.latest_viable_file, _ = self.get_latest_viable_spreadsheet(
            self.output_dir,
            min_bytes=self.min_viable_report_bytes,
        )
        logging.info(f"Latest spreadsheet: {self.latest_file}")
        if self.latest_file and self.latest_file != self.latest_viable_file:
            try:
                file_size = self.latest_file.stat().st_size
            except OSError:
                file_size = 0
            logging.warning(
                "Latest spreadsheet %s (%s bytes) failed viability guard; using %s as merge source.",
                self.latest_file,
                file_size,
                self.latest_viable_file,
            )
        
        final_records = []
        put_col = None
        call_col = None
        
        if self.latest_viable_file:
            df_existing = pd.read_excel(self.latest_viable_file)
            if "Last Update" not in df_existing.columns:
                df_existing["Last Update"] = None
            
            missing_or_old = self.get_missing_or_outdated_tickers(df_existing)
            
            # Start with existing records converted to list of dicts
            existing_records = df_existing.to_dict(orient='records')
            
            if not missing_or_old:
                logging.info("All tickers are up to date.")
                final_records = existing_records
                put_col = df_existing.columns.get_loc("Annual Yield Put Prem") + 1
                call_col = df_existing.columns.get_loc("Annual Yield Call Prem") + 1
            else:
                tickers_to_fetch = missing_or_old
                logging.info(f"Fetching data for {len(tickers_to_fetch)} tickers: {tickers_to_fetch}")
                
                # Filter out records that are about to be updated
                # We keep only records for tickers NOT in tickers_to_fetch
                preserved_records = [
                    r for r in existing_records 
                    if r.get("Ticker") not in tickers_to_fetch
                ]
                
                fetched_records = self.fetch_data(tickers_to_fetch)
                logging.info(f"fetched: {len(fetched_records)} records")
                
                # Combine preserved existing records with new fetched records
                final_records = preserved_records + fetched_records
                
        else:
            logging.info("No existing spreadsheet found. Fetching all data.")
            tickers_to_fetch = self.tickers
            final_records = self.fetch_data(tickers_to_fetch)

        # Create DataFrame from final combined list
        df = pd.DataFrame(final_records)
        if df.empty:
             df = pd.DataFrame(
                columns=["Ticker", "Annual Yield Put Prem", "Annual Yield Call Prem", "Last Update"]
            )
            
        # Deduplicate: Keep only the latest record for each Ticker
        if "Last Update" in df.columns and "Ticker" in df.columns:
            # Ensure Last Update is datetime for sorting
            df["Last Update"] = pd.to_datetime(df["Last Update"], errors='coerce')
            df = df.sort_values(by=["Last Update", "Ticker"], ascending=[False, True])
            initial_count = len(df)
            df = df.drop_duplicates(subset=["Ticker"], keep="first")
            final_count = len(df)
            if initial_count != final_count:
                logging.info(f"Removed {initial_count - final_count} duplicate records. Keeping {final_count} unique tickers.")
            # Convert Last Update back to string for consistency/Excel if needed, though datetime is fine.
            # But the current format is string "%Y-%m-%d %H:%M:%S".
            df["Last Update"] = df["Last Update"].dt.strftime("%Y-%m-%d %H:%M:%S")

        # Ensure we have column indices for formatting
        # If we didn't get them from existing DF, calculate them now
        # Note: add_ratio_column handles adding columns if missing
        df, put_col, call_col = self.add_ratio_column(df)

        if self.is_suspicious_record_count(len(df)):
            raise RuntimeError(
                f"Suspicious stock-analysis output ({len(df)} records for {len(self.tickers)} tickers). "
                "Skipping save to avoid overwriting a healthy report with truncated data."
            )
            
        self.save_to_excel(df, put_col, call_col)
        self.upsert_to_mongo(df)
        
        # Auto-export backup to JSON for git versioning
        try:
            logging.info("Auto-exporting MongoDB data to JSON backup...")
            export_data()
        except Exception as e:
            logging.error(f"Failed to auto-export JSON backup: {e}")
        
        logging.info(f"Spreadsheet generated: {self.filename}")

import argparse

def main():
    parser = argparse.ArgumentParser(description="Stock Live Comparison")
    parser.add_argument('--highlight-threshold', type=float, default=0.05, help='Highlight threshold as a decimal (default 0.05 for 5%)')
    args = parser.parse_args()

    StockLiveComparison.setup_logging()
    tickers = StockLiveComparison.get_default_tickers()
    logging.info(f"Processing {tickers} tickers...")

    comp = StockLiveComparison(tickers, highlight_threshold=args.highlight_threshold)
    comp.run()

if __name__ == "__main__":
    main()
